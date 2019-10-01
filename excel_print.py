from openpyxl import load_workbook
import time_my
import re
'''
def poisk(group):
	wb = load_workbook('./inputin.xlsx')
	sol=''
	sheet = wb["Table 1"]
	for i in range(1,sheet.max_column):
		s=chr(i+64)+'2'
		if str(sheet[s].value)==group:
			sol=s
	return(sol)
'''

def izm(ids, faculties, courses, groups):
	wb = load_workbook('./id.xlsx')
	sheet = wb["Sheet1"]
	x=0
	ids=str(ids)
	for i in range(1,sheet.max_row+1):
		if str(sheet['A'+str(i)].value)==ids:
			x=1
			sheet['B'+str(i)].value=faculties
			sheet['C'+str(i)].value=courses
			sheet['D'+str(i)].value=groups
			break
	if x==0:
		s1='A'+str(sheet.max_row+1)
		s2='B'+str(sheet.max_row+1)
		s3='C'+str(sheet.max_row+1)
		s4='D'+str(sheet.max_row+1)
		sheet[s1].value=ids
		sheet[s2].value=faculties
		sheet[s3].value=courses
		sheet[s4].value=groups			
	wb.save('./id.xlsx')

def output(ids):
	wb = load_workbook('./id.xlsx')
	sheet = wb["Sheet1"]
	sol=['','','']
	for i in range(1,sheet.max_row+1):
		if str(sheet['A'+str(i)].value)==ids:
			sol[0]=sheet['B'+str(i)].value
			sol[1]=sheet['C'+str(i)].value
			sol[2]=sheet['D'+str(i)].value
			break
	return(sol)		

def groups(fac, cours):
	groups=[]
	if int(cours)<5:
		if fac=="эконом" or fac=="строительный":
			s='./Расписание/'+cours+'econom_stroitel.xlsx'
			wb = load_workbook(s)
			sheet = wb['Page1']	
			if fac=="строительный":
				for i in range(3,sheet.max_column+1):
					if i<26:
						adr=chr(i+64)+'2'
					else:
						adr='A'+chr(i+38)+'2'
					if sheet[adr].value[:3]!='ИАС':
						groups.append(sheet[adr].value)
					else:
						return(groups)
						break
			else:
				i=sheet.max_column
				if i<27:
					adr=chr(i+64)+'2'
				else:
					adr='A'+chr(i+38)+'2'
				while sheet[adr].value[:3]!='ИАС': 
					if i<27:
						adr=chr(i+64)+'2'
					else:
						adr='A'+chr(i+38)+'2'
					groups.append(sheet[adr].value)
					i=i-1
				if groups[-1][-1]=='2':
					groups.append(groups[-1])
					groups[-1]=groups[-1].replace('2','1')
				groups.reverse()
				return(groups)
		elif fac=="горный" or fac=="фпмс":
			s='./Расписание/'+cours+'gor_fpms.xlsx'
			wb = load_workbook(s)
			sheet = wb['Page1']
			if fac=="горный":
				for i in range(3,sheet.max_column+1):
					if i<27:
						adr=chr(i+64)+'2'
					else:
						adr='A'+chr(i+38)+'2'
					if sheet[adr].value[:3]!='АПГ':
						groups.append(sheet[adr].value)
					else:
						return(groups)
						break
			else:
				i=sheet.max_column
				if i<27:
					adr=chr(i+64)+'2'
				else:
					adr='A'+chr(i+38)+'2'
				while sheet[adr].value[:3]!='АПГ': 
					if i<27:
						adr=chr(i+64)+'2'
					else:
						adr='A'+chr(i+38)+'2'
					groups.append(sheet[adr].value)
					i=i-1
				groups.reverse()
				return(groups)
		elif fac=="нефтегаз" or fac=="грф":
			s='./Расписание/'+cours+'neft_grf.xlsx'
			wb = load_workbook(s)
			sheet = wb['Page1']
			if fac=="грф":
				for i in range(3,sheet.max_column+1):
					if i<27:
						adr=chr(i+64)+'2'
					else:
						adr='A'+chr(i+38)+'2'
					if sheet[adr].value[:2]=='РФ':
						groups.append(sheet[adr].value)
						break
					else:
						groups.append(sheet[adr].value)
				return(groups)
			else:
				i=sheet.max_column
				if i<27:
					adr=chr(i+64)+'2'
				else:
					adr='A'+chr(i+38)+'2'	
				while sheet[adr].value[:2]!='РФ': 
					groups.append(sheet[adr].value)
					i=i-1
					if i<27:
						adr=chr(i+64)+'2'
					else:
						adr='A'+chr(i+38)+'2'
				groups.reverse()
				return(groups)
		else:
			s='./Расписание/'+cours+'emf.xlsx'
			wb = load_workbook(s)
			sheet = wb['Page1']
			for i in range(3,sheet.max_column+1):
				if i<27:
					adr=chr(i+64)+'2'
				else:
					adr='A'+chr(i+38)+'2'
				groups.append(sheet[adr].value)
			return(groups)
	else:
		if cours=='5':
			if fac=='грф':
				groups=['ГНГ-15', 'МГП-15', 'РГГ-15', 'РГИ-15', 'РМ-15', 'РФ-15', 'РФС-15']
			elif fac=='горный':
				groups=['БТС-15', 'ВД-15', 'ИЗС-15-1', 'ИЗС-15-2', 'ТО-15', 'ТПП-15', 'ТПР-15']
			elif fac=='строительный':
				groups=['АГС-15', 'ГГ-15-1', 'ГГ-15-2', 'ГС-15-1', 'ГС-15-2', 'ИГ-15-1', 'ИГ-15-2']	
			else:
				groups=['ГМ-15', 'ГТС-15', 'ЭРС-15-1', 'ЭРС-15-2', 'ОП-15', 'РТ-15 ']
		else:
			if fac=='горный':
				groups=['БТС-14', 'ВД-14', 'ИЗС-14-1', 'ИЗС-14-2', 'ТО-14', 'ТПП-14', 'ТПР-14']
			else:
				groups=['ГГ-14-1', 'ГГ-14-2', 'ГС-14-1', 'ГС-14-2', 'ГМ-14', 'ГТС-14', 'ЭРС-14']
		return(groups)


def next_les(fac, cours, group):
	dn={1:'ПОНЕДЕЛЬНИК',2:'ВТОРНИК',3:'СРЕДА',4:'ЧЕТВЕРГ',5:'ПЯТНИЦА',6:'СУББОТА', 7:'ВОСКРЕСЕНЬЕ'}
	y=0
	z=0
	q=0
	time=time_my.actual_time()     #Не забудь долбоеб
	para=None
	if int(cours)<5:
		if fac=="эконом" or fac=="строительный":
			s='./Расписание/'+cours+'econom_stroitel.xlsx'
		elif fac=="горный" or fac=="фпмс":
			s='./Расписание/'+cours+'gor_fpms.xlsx'
		elif fac=="нефтегаз" or fac=="грф":
			s='./Расписание/'+cours+'neft_grf.xlsx'
		else:
			s='./Расписание/'+cours+'emf.xlsx'
	else:
		if int(cours)==5:
			if fac=='грф' or fac=='нефтегаз' or fac=='фпмс' or fac=='эмф':
				s='./Расписание/'+cours+'grf_neft_fpms_emf.xlsx'
			else:
				s='./Расписание/'+cours+'gor_stroitel.xlsx'
		else:
			s=s='./Расписание/'+cours+'gor_fpms_stroitel.xlsx'
	wb = load_workbook(s)
	sheet = wb['Page1']
	for i in range(3,sheet.max_column+1):
		if i<27:
			adr=chr(i+64)+'2'
		else:
			adr='A'+chr(i+38)+'2'
		if sheet[adr].value==group:
			break
	sol=adr[0]
	n=time_my.day() #Не забудь долбоеб
	for i in range(3,sheet.max_row+1):
		adr='A'+str(i)
		if sheet[adr].value==dn[n]:
			if len(adr)>2:
				x=int(adr[1])*10+int(adr[2])
				break
			else:
				x=int(adr[1])
				break
	if n==7:
		n=0
	for i in range(3,sheet.max_row+1):
		adr='A'+str(i)
		if sheet[adr].value==dn[(n+1)]:
			if len(adr)>2:
				y=int(adr[1])*10+int(adr[2])
				break
			else:
				y=int(adr[1])
				break
	if time<8.5:
		s=0
	elif 8.5<=time<10.35:
		s=1
	elif 10.35<=time<12.35:
		s=2
	elif 12.35<=time<14.15:
		s=3	
	elif 14.15<=time<15.45:
		s=4	
	elif 15.45<=time<17.20:
		s=5
	elif 17.20<=time<19.00:
		s=5
	elif 19.00<=time<20.30:
		s=6
	elif 20.30<=time<22.00:
		s=7
	else:
		s=8
	#Ищем пару в этот же день Нет пары
	if (x+s)<y:
		for i in range(x+s,y):
				adr=sol+str(i)
				if nedelya(str(sheet[adr].value),q)!='Нет пары':
					para=nedelya(str(sheet[adr].value))
					break
	#Если не нашли, то идем дальше
	if para==None:
		for i in range(y,sheet.max_row+1):
			adr=sol+str(i)
			if nedelya(str(sheet[adr].value),q)!='Нет пары':
				para=nedelya(str(sheet[adr].value))
				z=1
				break
		if i==sheet.max_row and sheet[adr].value==None:
			q=1
			for i in range(3,sheet.max_row+1):
				adr=sol+str(i)
				if nedelya(str(sheet[adr].value),q)!='Нет пары':
					para=nedelya(str(sheet[adr].value),q)
					z=1
					break
	if z==1:
		para='Сегодня пар нет, следующая пара \n '+ para		
	return(para)	


def next_day(fac, cours, group):
	dn={1:'ПОНЕДЕЛЬНИК',2:'ВТОРНИК',3:'СРЕДА',4:'ЧЕТВЕРГ',5:'ПЯТНИЦА',6:'СУББОТА', 7:'ВОСКРЕСЕНЬЕ'}
	y=0
	s=''
	e=1
	z=''
	para=None
	#print(cours)
	if int(cours)<5:
		if fac=="эконом" or fac=="строительный":
			s='./Расписание/'+cours+'econom_stroitel.xlsx'
		elif fac=="горный" or fac=="фпмс":
			s='./Расписание/'+cours+'gor_fpms.xlsx'
		elif fac=="нефтегаз" or fac=="грф":
			s='./Расписание/'+cours+'neft_grf.xlsx'
		else:
			s='./Расписание/'+cours+'emf.xlsx'
	else:
		if int(cours)==5:
			if fac=='грф' or fac=='нефтегаз' or fac=='фпмс' or fac=='эмф':
				s='./Расписание/'+cours+'grf_neft_fpms_emf.xlsx'
			else:
				s='./Расписание/'+cours+'gor_stroitel.xlsx'
		else:
			s=s='./Расписание/'+cours+'gor_fpms_stroitel.xlsx'
	wb = load_workbook(s)
	sheet = wb['Page1']		
	for i in range(3,sheet.max_column+1):
		if i<27:
			adr=chr(i+64)+'2'
		else:
			adr='A'+chr(i+38)+'2'
		if sheet[adr].value==group:
			break
	sol=adr[0]
	n=time_my.day()+1
	if n==8:
		n=1
	for i in range(3,sheet.max_row+1):
		adr='A'+str(i)
		if sheet[adr].value==dn[n]:
			if len(adr)>2:
				x=int(adr[1])*10+int(adr[2])
				break
			else:
				x=int(adr[1])
				break
	for i in range(3,sheet.max_row+1):
		adr='A'+str(i)
		if sheet[adr].value==dn[n+1]:
			if len(adr)>2:
				y=int(adr[1])*10+int(adr[2])
				break
			else:
				y=int(adr[1])
				break
	for i in range(x,y):
		adr=sol+str(i)
		if sheet[adr].value!='None':
			z=z+'Пара '+str(e)+': '+nedelya(str(sheet[adr].value))+'\n'
		e=e+1
	if z=='':
		z='Завтра пар нет'
	return(z)

def to_day(fac, cours, group):
	dn={1:'ПОНЕДЕЛЬНИК',2:'ВТОРНИК',3:'СРЕДА',4:'ЧЕТВЕРГ',5:'ПЯТНИЦА',6:'СУББОТА', 7:'ВОСКРЕСЕНЬЕ'}
	y=0
	s=''
	e=1
	z=''
	para=None
	#print(cours)
	if int(cours)<5:
		if fac=="эконом" or fac=="строительный":
			s='./Расписание/'+cours+'econom_stroitel.xlsx'
		elif fac=="горный" or fac=="фпмс":
			s='./Расписание/'+cours+'gor_fpms.xlsx'
		elif fac=="нефтегаз" or fac=="грф":
			s='./Расписание/'+cours+'neft_grf.xlsx'
		else:
			s='./Расписание/'+cours+'emf.xlsx'
	else:
		if int(cours)==5:
			if fac=='грф' or fac=='нефтегаз' or fac=='фпмс' or fac=='эмф':
				s='./Расписание/'+cours+'grf_neft_fpms_emf.xlsx'
			else:
				s='./Расписание/'+cours+'gor_stroitel.xlsx'
		else:
			s=s='./Расписание/'+cours+'gor_fpms_stroitel.xlsx'
	wb = load_workbook(s)
	sheet = wb['Page1']		
	for i in range(3,sheet.max_column+1):
		if i<27:
			adr=chr(i+64)+'2'
		else:
			adr='A'+chr(i+38)+'2'
		if sheet[adr].value==group:
			break
	sol=adr[0]
	n=time_my.day()
	if n==8:
		n=1
	for i in range(3,sheet.max_row+1):
		adr='A'+str(i)
		if sheet[adr].value==dn[n]:
			if len(adr)>2:
				x=int(adr[1])*10+int(adr[2])
				break
			else:
				x=int(adr[1])
				break
	for i in range(3,sheet.max_row+1):
		adr='A'+str(i)
		if sheet[adr].value==dn[n+1]:
			if len(adr)>2:
				y=int(adr[1])*10+int(adr[2])
				break
			else:
				y=int(adr[1])
				break
	for i in range(x,y):
		adr=sol+str(i)
		if sheet[adr].value!='None':
			z=z+'Пара '+str(e)+': '+nedelya(str(sheet[adr].value))+'\n'
		e=e+1
	if z=='':
		z='Сегодня пар нет'
	return(z)

def nedelya(s,z=0):
	x=0
	s=re.sub('_','',s)
	s=' '.join(s.split())
	if time_my.day()==7 or z==1:
		x=int(time_my.week_number())+1
	else:
		x=int(time_my.week_number())
	if s[0]=='I':
		if x%2==1:
			try:
				s=s[:s.rindex('II')]
			except ValueError:
				pass
		else:
			try:
				s=s[s.rindex('II'):]
			except ValueError:
				pass
	if s=='' or s=='None':
		s='Нет пары'
	return(s)
		
#n=output('366145950')
#print(n[0], n[1], n[2])
#print(next_les('нефтегаз', "3", "НД-17-2"))
#print(time_my.actual_time())
#if actual_time>10.25 and actual_time<12.05
